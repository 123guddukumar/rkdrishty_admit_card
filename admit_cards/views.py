import os
import io
import zipfile
import threading
from datetime import datetime, date
from concurrent.futures import ThreadPoolExecutor
from openpyxl import load_workbook, Workbook
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse, FileResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.db.models import Q
from django.contrib import messages
from django.conf import settings
from .models import Student, ActivityLog, BulkTask
from .generator import generate_admit_card_files

# --- Background Task Worker ---
# A thread-safe executor for card generation in the background
executor = ThreadPoolExecutor(max_workers=4)

def log_activity(action, user, details):
    ActivityLog.objects.create(
        action=action,
        user=user if user and user.is_authenticated else None,
        details=details
    )

def bulk_generate_cards_worker(task_id, student_ids=None):
    """
    Background worker that runs card generation in a thread pool.
    """
    try:
        task = BulkTask.objects.get(id=task_id)
        task.status = 'RUNNING'
        task.save()

        if student_ids:
            students = Student.objects.filter(id__in=student_ids)
        else:
            students = Student.objects.filter(is_generated=False)

        task.total_records = students.count()
        task.save()

        if task.total_records == 0:
            task.status = 'COMPLETED'
            task.save()
            return

        processed = 0
        def process_single_student(student):
            nonlocal processed
            try:
                # 1. Generate files
                jpg_file, pdf_file = generate_admit_card_files(student)
                
                # 2. Save files to model
                student.card_jpg.save(jpg_file.name, jpg_file, save=False)
                student.card_pdf.save(pdf_file.name, pdf_file, save=False)
                student.is_generated = True
                student.save()
                
                processed += 1
                task.processed_records = processed
                task.save()
            except Exception as e:
                print(f"Error generating card for {student.name}: {e}")

        # Execute using thread pool for optimal performance (GIL released by Pillow)
        list(executor.map(process_single_student, students))

        task.status = 'COMPLETED'
        task.save()
        log_activity("Bulk Card Generation Completed", None, f"Generated {processed} cards successfully.")
    except Exception as e:
        task.status = 'FAILED'
        task.error_message = str(e)
        task.save()

def bulk_zip_worker(task_id):
    """
    Background worker that zips all generated cards.
    """
    try:
        task = BulkTask.objects.get(id=task_id)
        task.status = 'RUNNING'
        task.save()

        students = Student.objects.filter(is_generated=True)
        task.total_records = students.count()
        task.save()

        zip_dir = os.path.join(settings.MEDIA_ROOT, 'generated_cards')
        os.makedirs(zip_dir, exist_ok=True)
        zip_path = os.path.join(zip_dir, 'cards.zip')

        processed = 0
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for s in students:
                if s.card_jpg and os.path.exists(s.card_jpg.path):
                    zip_file.write(s.card_jpg.path, os.path.basename(s.card_jpg.path))
                if s.card_pdf and os.path.exists(s.card_pdf.path):
                    zip_file.write(s.card_pdf.path, os.path.basename(s.card_pdf.path))
                processed += 1
                task.processed_records = processed
                task.save()

        task.status = 'COMPLETED'
        task.save()
        log_activity("ZIP Bulk Export Completed", None, f"Packed {processed} cards into cards.zip.")
    except Exception as e:
        task.status = 'FAILED'
        task.error_message = str(e)
        task.save()

# --- Admin Dashboard Controllers ---

def dashboard(request):
    # Statistics
    total_students = Student.objects.count()
    generated_count = Student.objects.filter(is_generated=True).count()
    pending_count = Student.objects.filter(is_generated=False).count()
    
    # Generated today (local timezone)
    today = timezone.localtime(timezone.now()).date()
    generated_today = Student.objects.filter(is_generated=True, updated_at__date=today).count()

    # Search and filters
    query = request.GET.get('q', '').strip()
    class_filter = request.GET.get('class', '').strip()
    batch_filter = request.GET.get('batch', '').strip()
    status_filter = request.GET.get('status', '').strip()

    students_list = Student.objects.all().order_by('-created_at')

    if query:
        students_list = students_list.filter(
            Q(name__icontains=query) |
            Q(registration_number__icontains=query) |
            Q(mobile_number__icontains=query) |
            Q(aadhaar_number__icontains=query)
        )
    if class_filter:
        students_list = students_list.filter(student_class=class_filter)
    if batch_filter:
        students_list = students_list.filter(batch=batch_filter)
    if status_filter:
        is_gen = (status_filter == 'generated')
        students_list = students_list.filter(is_generated=is_gen)

    # Unique classes and batches for filters
    classes = Student.objects.values_list('student_class', flat=True).distinct()
    batches = Student.objects.values_list('batch', flat=True).distinct()

    # Running background tasks
    running_tasks = BulkTask.objects.filter(status__in=['PENDING', 'RUNNING'])

    # Activity logs
    logs = ActivityLog.objects.all().order_by('-timestamp')[:10]

    context = {
        'total_students': total_students,
        'generated_count': generated_count,
        'pending_count': pending_count,
        'generated_today': generated_today,
        'students': students_list,
        'classes': classes,
        'batches': batches,
        'logs': logs,
        'running_tasks': running_tasks,
        'query': query,
        'class_filter': class_filter,
        'batch_filter': batch_filter,
        'status_filter': status_filter,
    }
    return render(request, 'admit_cards/dashboard.html', context)

def parse_date_robust(date_str):
    if not date_str:
        return None
    date_str = date_str.strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Invalid date format: '{date_str}'. Please use YYYY-MM-DD or DD/MM/YYYY.")

def _is_ajax(request):
    return request.headers.get('X-Requested-With') == 'XMLHttpRequest'

def add_student(request):
    if request.method == 'POST':
        try:
            dob_str = request.POST.get('dob')
            dob_val = parse_date_robust(dob_str)
            
            student = Student.objects.create(
                name=request.POST.get('name'),
                father_name=request.POST.get('father_name'),
                mother_name=request.POST.get('mother_name'),
                student_class=request.POST.get('student_class'),
                batch=request.POST.get('batch', 'Batch 2026'),
                aadhaar_number=request.POST.get('aadhaar_number'),
                registration_number=request.POST.get('registration_number'),
                mobile_number=request.POST.get('mobile_number'),
                dob=dob_val,
                gender=request.POST.get('gender'),
                address=request.POST.get('address'),
                photo=request.FILES.get('photo'),
                signature=request.FILES.get('signature'),
            )
            log_activity("Add Student", request.user, f"Added student {student.name}")
            if _is_ajax(request):
                return JsonResponse({'status': 'success', 'message': f"Student {student.name} added successfully!", 'redirect': '/'})
            messages.success(request, f"Student {student.name} added successfully!")
            return redirect('dashboard')
        except Exception as e:
            if _is_ajax(request):
                return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
            messages.error(request, f"Error adding student: {e}")
            
    return render(request, 'admit_cards/student_form.html', {'action': 'Add'})

def edit_student(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        try:
            student.name = request.POST.get('name')
            student.father_name = request.POST.get('father_name')
            student.mother_name = request.POST.get('mother_name')
            student.student_class = request.POST.get('student_class')
            student.batch = request.POST.get('batch', 'Batch 2026')
            student.aadhaar_number = request.POST.get('aadhaar_number')
            student.registration_number = request.POST.get('registration_number')
            student.mobile_number = request.POST.get('mobile_number')
            
            dob_str = request.POST.get('dob')
            if dob_str:
                student.dob = parse_date_robust(dob_str)
                
            student.gender = request.POST.get('gender')
            student.address = request.POST.get('address')
            
            if request.FILES.get('photo'):
                student.photo = request.FILES.get('photo')
            if request.FILES.get('signature'):
                student.signature = request.FILES.get('signature')
                
            # If details changed, clear previous card files
            student.is_generated = False
            student.save()
            
            log_activity("Edit Student", request.user, f"Updated student {student.name}")
            if _is_ajax(request):
                return JsonResponse({'status': 'success', 'message': f"Student {student.name} updated successfully!", 'redirect': '/'})
            messages.success(request, f"Student {student.name} updated successfully!")
            return redirect('dashboard')
        except Exception as e:
            if _is_ajax(request):
                return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
            messages.error(request, f"Error updating student: {e}")
            
    return render(request, 'admit_cards/student_form.html', {'student': student, 'action': 'Edit'})

def delete_student(request, pk):
    student = get_object_or_404(Student, pk=pk)
    name = student.name
    # Delete file fields from media folder as well
    if student.card_jpg:
        student.card_jpg.delete(save=False)
    if student.card_pdf:
        student.card_pdf.delete(save=False)
    if student.photo:
        student.photo.delete(save=False)
    if student.signature:
        student.signature.delete(save=False)
        
    student.delete()
    log_activity("Delete Student", request.user, f"Deleted student {name}")
    messages.success(request, f"Student {name} deleted successfully!")
    return redirect('dashboard')

# --- Excel Utility Controllers ---

def bulk_import_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        file = request.FILES['excel_file']
        try:
            wb = load_workbook(file, data_only=True)
            sheet = wb.active
            
            # Column headers search
            headers = [str(cell.value).strip().lower() for cell in sheet[1]]
            
            # Map required columns
            col_map = {
                'name': ['student name', 'name', 'studentname'],
                'father': ['father name', 'father\'s name', 'fathername', 'father'],
                'mother': ['mother name', 'mother\'s name', 'mothername', 'mother'],
                'class': ['class', 'course'],
                'dob': ['dob', 'date of birth', 'dateofbirth'],
                'gender': ['gender', 'sex'],
                'address': ['address', 'pata'],
                'registration': ['registration', 'registration number', 'registration no', 'reg_no', 'regno'],
                'aadhaar': ['aadhaar', 'aadhar', 'aadhaar number', 'aadhar no'],
                'mobile': ['mobile', 'mobile number', 'mobile no', 'phone'],
                'photo_file': ['photo filename', 'photo_filename', 'photo', 'photo file']
            }
            
            indices = {}
            for field, aliases in col_map.items():
                indices[field] = None
                for i, h in enumerate(headers):
                    if h in aliases:
                        indices[field] = i
                        break
            
            # Check for critical column mappings
            if indices['name'] is None or indices['registration'] is None:
                messages.error(request, "Excel parsing error: Could not find 'Student Name' or 'Registration' column headers.")
                return redirect('dashboard')

            imported = 0
            # Read rows starting from row 2
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Skip empty rows
                if not row or not row[indices['name']]:
                    continue
                
                reg_num = str(row[indices['registration']]).strip()
                if not reg_num:
                    continue
                
                # Format Date of Birth
                dob_val = row[indices['dob']]
                if isinstance(dob_val, str):
                    try:
                        dob_val = datetime.strptime(dob_val.strip(), '%Y-%m-%d').date()
                    except ValueError:
                        try:
                            dob_val = datetime.strptime(dob_val.strip(), '%d/%m/%Y').date()
                        except ValueError:
                            dob_val = date.today()
                elif isinstance(dob_val, datetime):
                    dob_val = dob_val.date()
                elif not isinstance(dob_val, date):
                    dob_val = date.today()

                # Process photo if matching filename exists in student_photos
                photo_name = row[indices['photo_file']] if indices['photo_file'] is not None else None
                photo_path = None
                if photo_name:
                    # Look inside media/student_photos for matching filename
                    photo_name = str(photo_name).strip()
                    potential_path = os.path.join('student_photos', photo_name)
                    full_potential_path = os.path.join(settings.MEDIA_ROOT, potential_path)
                    if os.path.exists(full_potential_path):
                        photo_path = potential_path

                # Create or Update student
                Student.objects.update_or_create(
                    registration_number=reg_num,
                    defaults={
                        'name': str(row[indices['name']]).strip(),
                        'father_name': str(row[indices['father']]).strip() if indices['father'] is not None else "",
                        'mother_name': str(row[indices['mother']]).strip() if indices['mother'] is not None else "",
                        'student_class': str(row[indices['class']]).strip() if indices['class'] is not None else "",
                        'batch': request.POST.get('batch', 'Batch 2026'),
                        'dob': dob_val,
                        'gender': str(row[indices['gender']]).strip().capitalize() if indices['gender'] is not None else "Male",
                        'address': str(row[indices['address']]).strip() if indices['address'] is not None else "",
                        'aadhaar_number': str(row[indices['aadhaar']]).strip() if indices['aadhaar'] is not None else "",
                        'mobile_number': str(row[indices['mobile']]).strip() if indices['mobile'] is not None else "",
                        'photo': photo_path,
                        'is_generated': False
                    }
                )
                imported += 1

            log_activity("Excel Import", request.user, f"Successfully imported {imported} students.")
            messages.success(request, f"Excel import completed! {imported} students loaded.")
        except Exception as e:
            messages.error(request, f"Excel parsing error: {e}")
            
    return redirect('dashboard')

def export_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Students Admit Cards Data"
    
    headers = [
        "Student Name", "Father Name", "Mother Name", "Class", "Batch",
        "DOB", "Gender", "Address", "Registration", "Aadhaar", "Mobile",
        "Photo Filename", "Is Card Generated"
    ]
    ws.append(headers)
    
    for s in Student.objects.all():
        ws.append([
            s.name, s.father_name, s.mother_name, s.student_class, s.batch,
            s.dob.strftime('%Y-%m-%d') if s.dob else "", s.gender, s.address,
            s.registration_number, s.aadhaar_number, s.mobile_number,
            os.path.basename(s.photo.name) if s.photo else "",
            "Yes" if s.is_generated else "No"
        ])
        
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=drishti_students.xlsx"
    wb.save(response)
    return response

# --- REST APIs ---

@csrf_exempt
def generate_card_api(request):
    """
    POST /api/generate-card/
    Generates a card for a single student.
    """
    if request.method == 'POST':
        student_id = request.POST.get('student_id')
        reg_num = request.POST.get('registration_number')
        
        student = None
        if student_id:
            student = get_object_or_404(Student, id=student_id)
        elif reg_num:
            student = get_object_or_404(Student, registration_number=reg_num)
            
        if not student:
            return JsonResponse({'status': 'error', 'message': 'Student ID or Registration Number required.'}, status=400)
            
        try:
            jpg_file, pdf_file = generate_admit_card_files(student)
            student.card_jpg.save(jpg_file.name, jpg_file, save=False)
            student.card_pdf.save(pdf_file.name, pdf_file, save=False)
            student.is_generated = True
            student.save()
            
            return JsonResponse({
                'status': 'success',
                'message': f"Admit card generated for {student.name}",
                'jpg_url': student.card_jpg.url,
                'pdf_url': student.card_pdf.url
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f"Generation failed: {e}"}, status=500)
            
    return JsonResponse({'status': 'error', 'message': 'POST request required.'}, status=405)

@csrf_exempt
def generate_all_api(request):
    """
    POST /api/generate-all/
    Triggers asynchronous card generation in a background thread.
    """
    if request.method == 'POST':
        task = BulkTask.objects.create(
            task_type='BULK_GENERATE',
            status='PENDING'
        )
        # Launch generation in background thread
        thread = threading.Thread(target=bulk_generate_cards_worker, args=(task.id,))
        thread.start()
        
        return JsonResponse({
            'status': 'success',
            'message': 'Bulk card generation started in background.',
            'task_id': task.id
        })
        
    return JsonResponse({'status': 'error', 'message': 'POST request required.'}, status=405)

def download_card_api(request):
    """
    GET /api/download-card/?registration_number=XYZ&format=pdf
    Serves card files directly.
    """
    reg_num = request.GET.get('registration_number')
    student_id = request.GET.get('student_id')
    file_format = request.GET.get('format', 'jpg').lower()
    
    student = None
    if student_id:
        student = get_object_or_404(Student, id=student_id)
    elif reg_num:
        student = get_object_or_404(Student, registration_number=reg_num)
        
    if not student:
        return HttpResponse("Student parameter missing or invalid.", status=400)
        
    # Trigger generation if not done
    if not student.is_generated or not student.card_jpg or not student.card_pdf:
        try:
            jpg_file, pdf_file = generate_admit_card_files(student)
            student.card_jpg.save(jpg_file.name, jpg_file, save=False)
            student.card_pdf.save(pdf_file.name, pdf_file, save=False)
            student.is_generated = True
            student.save()
        except Exception as e:
            return HttpResponse(f"Error generating card: {e}", status=500)
            
    file_path = student.card_pdf.path if file_format == 'pdf' else student.card_jpg.path
    if not os.path.exists(file_path):
        return HttpResponse("Admit card file not found on server.", status=404)
        
    content_type = "application/pdf" if file_format == 'pdf' else "image/jpeg"
    ext = "pdf" if file_format == 'pdf' else "jpg"
    filename = f"{student.name.replace(' ', '_')}_admit_card.{ext}"
    
    response = FileResponse(open(file_path, 'rb'), content_type=content_type)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

@csrf_exempt
def download_zip_api(request):
    """
    GET /api/download-zip/
    Serves cards.zip or triggers background compression.
    """
    zip_path = os.path.join(settings.MEDIA_ROOT, 'generated_cards', 'cards.zip')
    
    # Check if cards.zip exists
    if os.path.exists(zip_path):
        # Update activity log
        log_activity("Download ZIP", None, "Downloaded cards.zip archive.")
        response = FileResponse(open(zip_path, 'rb'), content_type="application/zip")
        response['Content-Disposition'] = 'attachment; filename="cards.zip"'
        return response
        
    # If not generated, trigger ZIP builder task
    task = BulkTask.objects.create(
        task_type='BULK_ZIP',
        status='PENDING'
    )
    thread = threading.Thread(target=bulk_zip_worker, args=(task.id,))
    thread.start()
    
    return JsonResponse({
        'status': 'triggered',
        'message': 'ZIP generation started in the background. Please wait, then query status or retry download.',
        'task_id': task.id
    })

def task_status_api(request):
    """
    GET /api/task-status/?task_id=X
    Returns progress stats.
    """
    task_id = request.GET.get('task_id')
    if not task_id:
        return JsonResponse({'status': 'error', 'message': 'task_id required.'}, status=400)
        
    task = get_object_or_404(BulkTask, id=task_id)
    return JsonResponse({
        'task_id': task.id,
        'task_type': task.task_type,
        'status': task.status,
        'total_records': task.total_records,
        'processed_records': task.processed_records,
        'error_message': task.error_message,
        'progress_percent': int(task.processed_records / task.total_records * 100) if task.total_records > 0 else 0
    })
